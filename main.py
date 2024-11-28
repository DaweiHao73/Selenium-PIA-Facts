import time
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from openpyxl import load_workbook
from private_data import url
from private_data import email_address
from private_data import account
from private_data import password


def wait_for_element(driver, by, value, timeout=30):
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((by, value))
        )
        return element
    except TimeoutException:
        print(f"等待元素 {value} 超時")
        return None


def download_image(url, file_name):
    try:
        response = requests.get(url)
        if response.status_code == 200:
            content_type = response.headers.get('Content-Type', '')
            if 'image/jpeg' in content_type:
                ext = '.jpg'
            elif 'image/png' in content_type:
                ext = '.png'
            else:
                ext = '.img'
            full_file_name = f"{file_name}{ext}"
            with open(full_file_name, 'wb') as file:
                file.write(response.content)
            print(f"圖片成功下載: {full_file_name}")
        else:
            print(f"下載圖片失敗。狀態碼: {response.status_code}")
    except Exception as e:
        print(f"下載圖片時發生錯誤：{str(e)}")


def read_article_numbers(file_path, sheet_name, column_name):
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb[sheet_name]
    
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    print("標頭行內容:", header_row)
    
    try:
        column_index = header_row.index(column_name)
    except ValueError:
        print(f"錯誤：在標頭行中找不到列名 '{column_name}'")
        print("可用的列名:", ", ".join(header_row))
        return []
    
    article_numbers = [row[column_index] for row in ws.iter_rows(min_row=2, values_only=True) if row[column_index]]
    
    return article_numbers


def find_and_download_image(driver, article_number):
    xpaths = [
        f"//img[contains(@class, 'aspect-ratio-image__image') and contains(@alt, '{article_number}')]",
        f"//img[contains(@class, 'aspect-ratio-image__image')]",
        f"//img[contains(@class, 'aspect-ratio-image__image') and contains(@src, '{article_number}')]"
    ]
    
    for xpath in xpaths:
        try:
            img_elements = driver.find_elements(By.XPATH, xpath)
            for img_element in img_elements:
                img_src = img_element.get_attribute('src')
                if img_src:
                    alt_text = img_element.get_attribute('alt')
                    file_name = f"{article_number}"
                    download_image(img_src, file_name)
                    print(f"成功下載貨號 {article_number} 的商品圖片")
                    return True
        except Exception as e:
            print(f"使用 XPath '{xpath}' 查找圖片時出錯: {str(e)}")
    
    print(f"無法找到貨號 {article_number} 的商品圖片")
    return False

def auto_login_search_and_download(url, email, account, password, article_numbers):
    driver = None
    try:
        chrome_options = Options()
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-gpu")
        # chrome_options.add_argument("--headless")
        
        driver = webdriver.Chrome(options=chrome_options)
        print("")
        driver.get(url)
        
        # 登入過程
        email_input = wait_for_element(driver, By.ID, "username")
        if email_input:
            email_input.clear()
            email_input.send_keys(email)
            print("確認輸入 Email 成功")
        else:
            raise Exception("無法找到電子郵件輸入欄位")
        
        # 修改後的繼續按鈕處理邏輯
        continue_button = None
        selectors = [
            (By.CLASS_NAME, "_button-login-id"),
            (By.CSS_SELECTOR, ".c0a486a03.c3a925026.cc4e2760d.cf0fbb154._button-login-id"),
            (By.XPATH, "//button[contains(text(), '繼續')]"),
            (By.XPATH, "//button[contains(@type, 'submit')]"),
            (By.XPATH, "//button[contains(text(), 'Continue')]")
        ]

        for selector_type, selector_value in selectors:
            try:
                continue_button = wait_for_element(driver, selector_type, selector_value, timeout=10)
                if continue_button and continue_button.is_displayed() and continue_button.is_enabled():
                    continue_button.click()
                    print("點擊continue button 成功")
                    break
            except Exception as e:
                print(f"嘗試選擇器 {selector_value} 失敗，嘗試下一個...")
                continue

        if not continue_button:
            raise Exception("無法找到繼續按鈕")

        # 等待頁面加載
        time.sleep(2)
        
        input_account = wait_for_element(driver, By.ID, "userNameInput")
        if input_account:
            input_account.send_keys(account)
            print("確認輸入 account 成功")
        else:
            raise Exception("無法找到帳號輸入欄位")
        
        input_password = wait_for_element(driver, By.ID, "passwordInput")
        if input_password:
            input_password.send_keys(password)
            print("確認輸入 password 成功")
        else:
            raise Exception("無法找到密碼輸入欄位")
        
        sign_in_btn = wait_for_element(driver, By.ID, "submitButton")
        if sign_in_btn:
            sign_in_btn.click()
            print("確認點擊sign_in_btn button 成功")
        else:
            raise Exception("無法找到登入按鈕")
        
        # 增加等待時間以確保頁面完全加載
        time.sleep(3)
        
        selector = wait_for_element(driver, By.ID, "unit-selector", timeout=60)
        if selector:
            select = Select(selector)
            select.select_by_value("zh-TW-TW")
            print("確認已點擊selector，並選取 zh-TW-TW")
        else:
            raise Exception("無法找到語言選擇器")
        
        time.sleep(5)
        
        for article_number in article_numbers:
            tries = 0
            while tries < 3:
                try:
                    search_input = wait_for_element(driver, By.ID, "search", timeout=20)
                    if search_input:
                        search_input.clear()
                        time.sleep(1)  # 加入短暫延遲確保清除完成
                        search_input.send_keys(article_number)
                        time.sleep(1)  # 加入短暫延遲確保輸入完成
                        search_input.send_keys(Keys.ENTER)
                        break
                    else:
                        time.sleep(2)
                except Exception as e:
                    print(f"搜尋嘗試 {tries + 1} 失敗: {str(e)}")
                    tries += 1
                    time.sleep(2)
            else:
                print(f"無法找到貨號 {article_number} 的搜尋欄位")
                continue
            
            # 增加頁面加載等待時間
            time.sleep(15)  # 增加等待時間，確保圖片完全加載
            
            if not find_and_download_image(driver, article_number):
                print(f"無法找到貨號 {article_number} 的圖片，繼續下一個貨號")
            
            # 返回主頁面，準備下一次搜索
            driver.get(url)
            time.sleep(7)  # 增加等待時間，確保頁面完全加載
        
        print("所有貨號的搜尋和圖片下載過程已完成")
    
    except Exception as e:
        print(f"發生錯誤：{str(e)}")
        import traceback
        print("完整的錯誤追蹤：")
        print(traceback.format_exc())
    
    finally:
        if driver:
            print("按Enter鍵關閉瀏覽器...")
            input()
            driver.quit()

# 主程式
if __name__ == "__main__":
    
    # 讀取 Excel 文件中的貨號
    excel_file_path = "file_1.xlsx"
    sheet_name = "Sheet1"   # 根據實際工作表名稱修改
    column_name = "art"     # 根據實際列名修改
    
    article_numbers = read_article_numbers(excel_file_path, sheet_name, column_name)
    
    auto_login_search_and_download(url, email_address, account, password, article_numbers)